import pandas as pd
import logging
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def save_to_excel(
    df: pd.DataFrame, 
    top_5: pd.DataFrame, 
    average_price: float, 
    highest_change: dict, 
    lowest_change: dict, 
    excel_file: str = 'crypto_tracker.xlsx'
) -> None:
    """
    Save cryptocurrency data to an Excel file with enhanced summary sheet.
    
    Args:
        df (DataFrame): Main cryptocurrency data
        top_5 (DataFrame): Top 5 cryptocurrencies
        average_price (float): Average price of top 50 cryptos
        highest_change (dict): Crypto with highest 24h change
        lowest_change (dict): Crypto with lowest 24h change
        excel_file (str): Path to Excel file
    """
    try:
        os.makedirs(os.path.dirname(excel_file) or '.', exist_ok=True)
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Live Crypto Data', index=False)
            workbook = writer.book
            main_sheet = writer.sheets['Live Crypto Data']
            
            for col in range(1, len(df.columns) + 1):
                col_letter = get_column_letter(col)
                main_sheet.column_dimensions[col_letter].width = 20
                
            top_5.to_excel(writer, sheet_name='Top 5 Cryptos', index=False)
            
            summary_data = {
                'Market Overview': [
                    'Timestamp',
                    'Total Cryptocurrencies Tracked',
                    'Average Cryptocurrency Price',
                    'Total Market Capitalization',
                    'Total 24h Trading Volume'
                ],
                'Value': [
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    len(df),
                    f'${average_price:.2f}',
                    f'${df["Market Cap"].sum():,.2f}',
                    f'${df["Trading Volume"].sum():,.2f}'
                ]
            }
            
            performance_data = {
                'Performance Metrics': [
                    'Best Performing Cryptocurrency',
                    'Highest 24h Price Change',
                    'Worst Performing Cryptocurrency',
                    'Lowest 24h Price Change'
                ],
                'Details': [
                    highest_change.get('Name', 'N/A'),
                    f"{highest_change.get('Price Change (%)', 'N/A')}%",
                    lowest_change.get('Name', 'N/A'),
                    f"{lowest_change.get('Price Change (%)', 'N/A')}%"
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            performance_df = pd.DataFrame(performance_data)
            
            summary_df.to_excel(writer, sheet_name='Market Summary', index=False)
            performance_df.to_excel(writer, sheet_name='Performance Highlights', index=False)
        
        logging.info(f"Data saved to {excel_file} successfully")
    
    except Exception as e:
        logging.error(f"Error saving to Excel: {e}")